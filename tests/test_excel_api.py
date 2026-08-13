from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app


client = TestClient(app)


def create_excel(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def sample_excel() -> bytes:
    return create_excel(
        ["商品名称", "售价", "成本", "运费", "佣金率"],
        [
            ["无线鼠标", 79, 35, 8, 0.05],
            ["手机壳", 19, 15, 5, 0.05],
            [None, 29, 10, 4, 0.05],
        ],
    )


def test_analyze_excel_with_success_and_error_rows() -> None:
    response = client.post(
        "/products/analyze-excel",
        files={
            "file": (
                "products.xlsx",
                sample_excel(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total_rows"] == 3
    assert body["success_count"] == 2
    assert body["error_count"] == 1
    assert body["results"][0]["analysis"]["profit"] == 32.05
    assert body["results"][1]["analysis"]["profit"] == -1.95
    assert body["results"][2]["status"] == "error"
    assert "商品名称" in body["results"][2]["error_reason"]


def test_analyze_excel_rejects_missing_header() -> None:
    content = create_excel(
        ["商品名称", "售价", "成本"],
        [["无线鼠标", 79, 35]],
    )
    response = client.post(
        "/products/analyze-excel",
        files={"file": ("products.xlsx", content)},
    )

    assert response.status_code == 400
    assert "缺少必要表头" in response.json()["detail"]


def test_analyze_excel_rejects_non_xlsx_file() -> None:
    response = client.post(
        "/products/analyze-excel",
        files={"file": ("products.csv", b"name,price")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "只支持 .xlsx 格式的 Excel 文件"


def test_export_excel_returns_workbook() -> None:
    response = client.post(
        "/products/analyze-excel/export",
        files={"file": ("products.xlsx", sample_excel())},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    worksheet = workbook["分析结果"]
    assert worksheet.cell(row=1, column=1).value == "原始行号"
    assert worksheet.cell(row=2, column=9).value == 32.05
    assert worksheet.cell(row=4, column=13).value == "失败"
    workbook.close()
