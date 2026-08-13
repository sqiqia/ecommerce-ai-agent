from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.product import ExcelAnalyzeResponse


REQUIRED_HEADERS = {
    "商品名称": "product_name",
    "售价": "sale_price",
    "成本": "cost_price",
    "运费": "shipping_fee",
    "佣金率": "commission_rate",
}

RESULT_HEADERS = [
    "原始行号",
    "商品名称",
    "售价",
    "成本",
    "运费",
    "佣金率",
    "平台佣金",
    "总成本",
    "利润",
    "利润率",
    "是否盈利",
    "经营建议",
    "数据状态",
    "错误原因",
]


class ExcelValidationError(ValueError):
    """上传文件不是项目要求的商品 Excel 时抛出。"""


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_cell(value: Any) -> float | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def read_product_rows(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    """读取第一张工作表，返回 Excel 原始行号和字段数据。"""

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ExcelValidationError("文件无法读取，请确认它是有效的 .xlsx 文件") from exc

    try:
        worksheet = workbook.active
        header_values = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if not header_values:
            raise ExcelValidationError("Excel 没有表头")

        header_map: dict[str, int] = {}
        duplicate_headers: set[str] = set()
        for index, value in enumerate(header_values):
            header = normalize_header(value)
            if not header:
                continue
            if header in header_map:
                duplicate_headers.add(header)
            header_map[header] = index

        if duplicate_headers:
            names = "、".join(sorted(duplicate_headers))
            raise ExcelValidationError(f"Excel 存在重复表头：{names}")

        missing_headers = [name for name in REQUIRED_HEADERS if name not in header_map]
        if missing_headers:
            names = "、".join(missing_headers)
            raise ExcelValidationError(f"Excel 缺少必要表头：{names}")

        rows: list[tuple[int, dict[str, Any]]] = []
        for source_row, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            product_data = {
                field_name: normalize_cell(row_values[header_map[header_name]])
                if header_map[header_name] < len(row_values)
                else None
                for header_name, field_name in REQUIRED_HEADERS.items()
            }

            if all(
                value is None or (isinstance(value, str) and not value.strip())
                for value in product_data.values()
            ):
                continue

            rows.append((source_row, product_data))

        if not rows:
            raise ExcelValidationError("Excel 中没有可处理的商品数据")

        return rows
    finally:
        workbook.close()


def build_result_workbook(batch: ExcelAnalyzeResponse) -> bytes:
    """把批量分析结果导出为新的 Excel 文件。"""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "分析结果"
    worksheet.append(RESULT_HEADERS)

    for result in batch.results:
        analysis = result.analysis
        worksheet.append(
            [
                result.source_row,
                result.product_name,
                result.sale_price,
                result.cost_price,
                result.shipping_fee,
                result.commission_rate,
                analysis.commission if analysis else None,
                analysis.total_cost if analysis else None,
                analysis.profit if analysis else None,
                analysis.profit_rate if analysis else None,
                "是" if analysis and analysis.profitable else "否" if analysis else None,
                analysis.advice if analysis else None,
                "成功" if result.status == "success" else "失败",
                result.error_reason,
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F766E")
    success_fill = PatternFill("solid", fgColor="DCFCE7")
    error_fill = PatternFill("solid", fgColor="FEE2E2")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_number, column=6).number_format = "0.00%"
        worksheet.cell(row=row_number, column=10).number_format = "0.00%"
        for column in range(3, 10):
            if column not in (6, 10):
                worksheet.cell(row=row_number, column=column).number_format = "0.00"
        status_cell = worksheet.cell(row=row_number, column=13)
        status_cell.fill = success_fill if status_cell.value == "成功" else error_fill

    column_widths = {
        "A": 10,
        "B": 18,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 34,
        "M": 12,
        "N": 36,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
